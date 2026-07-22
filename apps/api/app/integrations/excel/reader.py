from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook

SourceRole = Literal["live_actual", "anchor_schedule", "staff_schedule", "unknown"]


@dataclass(frozen=True)
class SheetScan:
    name: str
    dimension: str
    rows: int
    columns: int
    headers: tuple[str, ...]
    source_role: SourceRole


@dataclass(frozen=True)
class WorkbookScan:
    filename: str
    sheets: tuple[SheetScan, ...]


@dataclass(frozen=True)
class FixtureRecord:
    source_record_id: str
    source_role: SourceRole
    default_room_name: str | None
    raw_fields: dict[str, Any]
    payload_hash: str


def detect_source_role(headers: set[str]) -> SourceRole:
    if {"主播", "场控", "自动检查", "时间", "时段"}.issubset(headers):
        return "live_actual"
    if {"直播间", "月份", "时段", "1日"}.issubset(headers):
        return "anchor_schedule"
    if {"姓名", "岗位", "状态", "月份", "1日"}.issubset(headers):
        return "staff_schedule"
    return "unknown"


def scan_workbook(path: Path) -> WorkbookScan:
    workbook = load_workbook(path, read_only=False, data_only=False)
    sheets: list[SheetScan] = []
    for sheet in workbook.worksheets:
        headers = tuple(
            str(sheet.cell(1, column).value or "").strip()
            for column in range(1, sheet.max_column + 1)
        )
        sheets.append(
            SheetScan(
                name=sheet.title,
                dimension=sheet.calculate_dimension(),
                rows=sheet.max_row,
                columns=sheet.max_column,
                headers=headers,
                source_role=detect_source_role(set(headers)),
            )
        )
    workbook.close()
    return WorkbookScan(filename=path.name, sheets=tuple(sheets))


def iter_fixture_records(path: Path) -> list[FixtureRecord]:
    # The supplied workbook has an invalid streaming dimension cache (A1). Normal mode
    # reads the actual worksheet XML dimensions and remains safe because formulas are
    # never executed by openpyxl.
    workbook = load_workbook(path, read_only=False, data_only=False)
    records: list[FixtureRecord] = []
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        headers = tuple(str(value or "").strip() for value in next(rows))
        role = detect_source_role(set(headers))
        for row_number, values in enumerate(rows, start=2):
            raw_fields = {
                header: _json_safe(value)
                for header, value in zip(headers, values, strict=False)
                if header
            }
            if not any(value is not None and value != "" for value in raw_fields.values()):
                continue
            normalized = json.dumps(
                raw_fields, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str
            )
            records.append(
                FixtureRecord(
                    source_record_id=f"{sheet.title}:{row_number}",
                    source_role=role,
                    default_room_name=sheet.title if role == "live_actual" else None,
                    raw_fields=raw_fields,
                    payload_hash=hashlib.sha256(normalized.encode()).hexdigest(),
                )
            )
    workbook.close()
    return records


def _json_safe(value: Any) -> Any:
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)
