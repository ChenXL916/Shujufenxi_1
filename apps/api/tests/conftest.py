from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pytest
import yaml
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[3]


@dataclass(frozen=True)
class ExcelFixtureSet:
    live: Path
    schedule: Path
    paths: tuple[Path, Path]
    private: bool


def _create_synthetic_live_workbook(path: Path) -> None:
    seed = yaml.safe_load((ROOT / "config" / "metric_seed.yml").read_text(encoding="utf-8"))
    headers = [*seed["dimensions"], *(metric["field"] for metric in seed["metrics"])]
    workbook = Workbook()
    for index, room_name in enumerate(("E2E-测试直播间A", "E2E-测试直播间B")):
        sheet = workbook.active if index == 0 else workbook.create_sheet()
        sheet.title = room_name
        sheet.append(headers)
        for row_index, hour in enumerate((8, 9, 10), start=1):
            fields = {header: 1 for header in headers}
            fields.update(
                {
                    "主播": f"E2E主播{index + 1}",
                    "场控": f"E2E场控{index + 1}",
                    "月份": 7,
                    "自动检查": "错误" if row_index == 3 else "正确",
                    "时间": f"2026-07-08 {hour:02d}:30:00",
                    "时段": f"{hour:02d}-{hour + 1:02d}",
                    "时段消耗": 100 * row_index,
                    "时段整体成交金额": 200 * row_index,
                    "时段整体支付ROI": 2,
                }
            )
            sheet.append([fields[header] for header in headers])
    workbook.save(path)
    workbook.close()


def _create_synthetic_schedule_workbook(path: Path) -> None:
    days = [f"{day}日" for day in range(1, 32)]
    workbook = Workbook()
    staff = workbook.active
    staff.title = "直播部门排班表"
    staff.append(["姓名", "岗位", "状态", "月份", *days])
    staff.append(["E2E场控1", "场控", "在职", 7, *("08-17" for _ in days)])
    staff.append(["E2E场控2", "场控", "在职", 7, *("休息" for _ in days)])

    anchors = workbook.create_sheet("主播直播排班表")
    anchors.append(["直播间", "月份", "时段", *days])
    anchors.append(
        [
            "E2E-测试直播间A",
            7,
            "08-09",
            "断播",
            "E2E主播1+E2E主播2",
            *("E2E主播1" for _ in days[2:]),
        ]
    )
    anchors.append(["E2E-测试直播间B", 7, "09-10", *("E2E主播2" for _ in days)])
    workbook.save(path)
    workbook.close()


@pytest.fixture(scope="session")
def excel_fixture_set(tmp_path_factory: pytest.TempPathFactory) -> ExcelFixtureSet:
    private_paths = sorted((ROOT / "fixtures").glob("*.xlsx"))
    live = next((path for path in private_paths if path.stat().st_size > 100_000), None)
    schedule = next((path for path in private_paths if path.stat().st_size <= 100_000), None)
    if live is not None and schedule is not None:
        return ExcelFixtureSet(live=live, schedule=schedule, paths=(live, schedule), private=True)

    directory = tmp_path_factory.mktemp("synthetic_excel_fixtures")
    live = directory / "synthetic-live.xlsx"
    schedule = directory / "synthetic-schedule.xlsx"
    _create_synthetic_live_workbook(live)
    _create_synthetic_schedule_workbook(schedule)
    return ExcelFixtureSet(live=live, schedule=schedule, paths=(live, schedule), private=False)
