from datetime import date, time
from decimal import Decimal
from pathlib import Path

import pytest
import yaml
from openpyxl import load_workbook

from app.domain.cleaning import (
    invalid_live_record_reason,
    normalize_hour_slot,
    normalize_person_name,
    parse_business_datetime,
    parse_decimal,
    parse_shift,
    unpivot_anchor_schedule,
    unpivot_staff_schedule,
)

ROOT = Path(__file__).resolve().parents[3]


def test_time_parsing_and_hour_normalization() -> None:
    parsed = parse_business_datetime("2026-07-08T09:31:00+08:00")
    slot = normalize_hour_slot("0:00-1:00", parsed.date())
    late = normalize_hour_slot("23:00-24:00", date(2026, 7, 8))

    assert parsed.isoformat() == "2026-07-08T09:31:00+08:00"
    assert slot.label == "00-01"
    assert late.label == "23-24"
    assert late.end_at.date() == date(2026, 7, 9)
    with pytest.raises(ValueError, match="不是有效"):
        normalize_hour_slot("0:00-0:00", parsed.date())


def test_feishu_millisecond_and_excel_serial_date_parse() -> None:
    assert parse_business_datetime(1_752_000_000_000).tzinfo is not None
    assert parse_business_datetime(45_000).tzinfo is not None


def test_person_notes_prefix_and_combination_members() -> None:
    noted = normalize_person_name("Q-李昕（9.31开播）")
    combination = normalize_person_name("J-梦丽+菜菜")
    mentioned = normalize_person_name("@陈铭玉")

    assert noted is not None
    assert (noted.canonical, noted.base_name, noted.note) == ("Q-李昕", "李昕", "9.31开播")
    assert combination is not None
    assert combination.members == ("梦丽", "菜菜")
    assert mentioned is not None
    assert (mentioned.canonical, mentioned.base_name) == ("陈铭玉", "陈铭玉")
    assert normalize_person_name("断播") is None
    assert normalize_person_name("用于计算") is None


def test_decimal_and_invalid_rows() -> None:
    assert parse_decimal("9.68%") == Decimal("0.0968")
    assert parse_decimal("1,234.50") == Decimal("1234.50")
    assert parse_decimal("—") is None
    assert invalid_live_record_reason({"主播": "用于计算"}) == "用于计算"
    assert invalid_live_record_reason({"主播": "A", "自动检查": "错误"}) == "自动检查错误"
    assert (
        invalid_live_record_reason({"主播": "A", "自动检查": [{"text": "错误", "type": "text"}]})
        == "自动检查错误"
    )


def test_shift_cross_midnight_and_unconfigured_text() -> None:
    overnight = parse_shift("20-05")
    text_shift = parse_shift("晚班")

    assert (overnight.start, overnight.end, overnight.crosses_midnight) == (time(20), time(5), True)
    assert overnight.time_configured is True
    assert text_shift.time_configured is False


@pytest.mark.integration
def test_real_schedule_fixture_unpivots_both_tables(
    excel_fixture_set,  # type: ignore[no-untyped-def]
) -> None:
    workbook = load_workbook(excel_fixture_set.schedule, data_only=False)
    staff_sheet = workbook["直播部门排班表"]
    anchor_sheet = workbook["主播直播排班表"]
    staff_headers = [cell.value for cell in staff_sheet[1]]
    anchor_headers = [cell.value for cell in anchor_sheet[1]]
    staff_record = dict(zip(staff_headers, [cell.value for cell in staff_sheet[2]], strict=False))
    anchor_records = [
        dict(zip(anchor_headers, [cell.value for cell in row], strict=False))
        for row in anchor_sheet.iter_rows(min_row=2)
    ]

    assert len(unpivot_staff_schedule(staff_record, 2026)) == 31
    all_rows = [row for record in anchor_records for row in unpivot_anchor_schedule(record, 2026)]
    assert len(all_rows) == len(anchor_records) * 31
    assert any(row.schedule_status == "off_air" for row in all_rows)
    assert any(row.schedule_status == "combination" for row in all_rows)


@pytest.mark.integration
def test_metric_seed_covers_every_numeric_fixture_field(
    excel_fixture_set,  # type: ignore[no-untyped-def]
) -> None:
    workbook = load_workbook(excel_fixture_set.live, read_only=False, data_only=False)
    headers = {str(cell.value) for cell in workbook.worksheets[0][1]}
    seed = yaml.safe_load((ROOT / "config" / "metric_seed.yml").read_text(encoding="utf-8"))
    dimensions = set(seed["dimensions"])
    metric_fields = {metric["field"] for metric in seed["metrics"]}

    assert len(headers) == 52
    assert len(dimensions) == 6
    assert len(metric_fields) == 46
    assert headers == dimensions | metric_fields
