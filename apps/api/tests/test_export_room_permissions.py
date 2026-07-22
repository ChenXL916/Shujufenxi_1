from __future__ import annotations

import csv
import io
from datetime import date, datetime, timedelta
from uuid import UUID, uuid4
from zoneinfo import ZoneInfo

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.auth.dependencies import AccessScope, get_access_scope
from app.db.base import Base
from app.db.session import get_db
from app.main import app
from app.models.entities import HourlyFact, Room

SHANGHAI = ZoneInfo("Asia/Shanghai")


def _add_fact(session: Session, room_id: UUID, anchor: str, hour: int) -> None:
    start = datetime(2026, 7, 15, hour, tzinfo=SHANGHAI)
    session.add(
        HourlyFact(
            room_id=room_id,
            business_date=date(2026, 7, 15),
            year=2026,
            month=7,
            hour_slot=f"{hour:02d}-{hour + 1:02d}",
            hour_order=hour,
            hour_start_at=start,
            hour_end_at=start + timedelta(hours=1),
            latest_point_id=None,
            latest_observed_at=start,
            actual_anchor_canonical=anchor,
            actual_anchor_base_names=[anchor],
            actual_control_canonical="测试场控",
            planned_anchor_canonical=anchor,
            planned_anchor_base_names=[anchor],
            anchor_schedule_status="scheduled",
            anchor_match_status="matched",
            control_shift_name="测试班次",
            control_is_scheduled=True,
            control_is_rest=False,
            control_may_be_on_duty=True,
            data_status="complete",
        )
    )


def test_general_export_rejects_rooms_without_export_permission() -> None:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        allowed = Room(
            name="可导出直播间",
            brand=None,
            category=None,
            active=True,
            confirmed=True,
            source_aliases=[],
        )
        view_only = Room(
            name="仅查看直播间",
            brand=None,
            category=None,
            active=True,
            confirmed=True,
            source_aliases=[],
        )
        session.add_all([allowed, view_only])
        session.flush()
        _add_fact(session, allowed.id, "AUTHORIZED_EXPORT_MARKER", 8)
        _add_fact(session, view_only.id, "VIEW_ONLY_NO_EXPORT_MARKER", 9)
        session.commit()
        allowed_id = allowed.id
        view_only_id = view_only.id

    def override_db():  # type: ignore[no-untyped-def]
        with Session(engine) as session:
            yield session

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[get_access_scope] = lambda: AccessScope(
        user_id=uuid4(),
        role="viewer",
        room_ids=frozenset({allowed_id, view_only_id}),
        can_export=True,
        export_room_ids=frozenset({allowed_id}),
        permission_codes=frozenset({"dashboard.view", "dashboard.export"}),
    )
    client = TestClient(app)
    try:
        mixed = client.post(
            "/api/v1/exports",
            params={
                "file_format": "csv",
                "start_date": "2026-07-15",
                "end_date": "2026-07-15",
            },
        )
        denied = client.post(
            "/api/v1/exports",
            params={
                "file_format": "csv",
                "start_date": "2026-07-15",
                "end_date": "2026-07-15",
                "room_ids": str(view_only_id),
            },
        )
        permitted = client.post(
            "/api/v1/exports",
            params={
                "file_format": "csv",
                "start_date": "2026-07-15",
                "end_date": "2026-07-15",
                "room_ids": str(allowed_id),
            },
        )
        permitted_xlsx = client.post(
            "/api/v1/exports",
            params={
                "file_format": "xlsx",
                "start_date": "2026-07-15",
                "end_date": "2026-07-15",
                "room_ids": str(allowed_id),
            },
        )
    finally:
        app.dependency_overrides.clear()
        engine.dispose()

    assert mixed.status_code == 403
    assert denied.status_code == 403
    assert permitted.status_code == 200
    assert permitted_xlsx.status_code == 200
    csv_text = permitted.content.decode("utf-8-sig")
    assert "AUTHORIZED_EXPORT_MARKER" in csv_text
    assert "VIEW_ONLY_NO_EXPORT_MARKER" not in csv_text
    csv_rows = list(csv.reader(io.StringIO(csv_text)))
    assert csv_rows[0][:2] == ["直播间ID", "直播间"]
    assert csv_rows[1][:2] == [str(allowed_id), "可导出直播间"]

    workbook = load_workbook(io.BytesIO(permitted_xlsx.content), read_only=True, data_only=True)
    sheet = workbook.active
    assert sheet is not None
    xlsx_rows = list(sheet.iter_rows(min_row=1, max_row=2, values_only=True))
    workbook.close()
    assert list(xlsx_rows[0][:2]) == ["直播间ID", "直播间"]
    assert list(xlsx_rows[1][:2]) == [str(allowed_id), "可导出直播间"]
